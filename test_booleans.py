# This script checks the validity of the surface mapping computed with our surface mapping check.
# Our run_surface_mapping.py script creates the folders "vol_in" and "srf_maps". 
# Put this script inside the "vol_in" folder.
# For each "<filename>.mesh" in the "vol_in" folder, the script checks the mapping in the "srf_maps" folder  
# (with name <filename>_smap.txt, as produced from our run_surface_mapping.py script).


import smtplib
from email.message import EmailMessage
import os
import subprocess
import time
import openpyxl
import re
from openpyxl import load_workbook


def sorted_alphanumeric(data):
    convert = lambda text: int(text) if text.isdigit() else text.lower()
    alphanum_key = lambda key: [convert(c) for c in re.split('([0-9]+)', key) ]
    return sorted(data, key=alphanum_key)



def completion_bar(iterations, total):
    # Calculate the completion ratio
    ratio = iterations / total

    # Determine the length of the progress bar
    bar_length = 50
    progress = int(bar_length * ratio)

    # Create the completion bar
    bar = "[" + "#" * progress + "-" * (bar_length - progress) + "]"

    return bar

def clean_folder(folder_path):
    # Iterate over the contents of the folder
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        # Check if the path is a file or directory
        if os.path.isfile(file_path) and ".xlsx" not in file_path:
            # If it's a file but not a xlsx file, remove it 
            os.remove(file_path)
        elif os.path.isdir(file_path):
            # If it's a directory, recursively clean it
            clean_folder(file_path)
            # Remove the empty directory
            os.rmdir(file_path)


def compile_and_run_cpp(mesh_file, mesh_file_rotated, path_folder_filename, operation, excel_file, mode="release"):
    executable = os.path.join("cmake-build-"+ mode, 'mesh_booleans')
    #executable = '/Users/michele/Documents/GitHub/FastAndRobustMeshArrangements/history-results/old_version_results/mesh_arrangement'
    #measure the time it takes to run the process
    start_time_float = time.time()

    #remove the path from the mesh_file
    mesh_file_output = mesh_file.replace(path_folder_filename + "/", "") + ".obj"
    
    path_folder_output = "/Users/michele/Documents/GitHub/InteractiveAndRobustMeshBooleans/results_tests/meshes"
    run_process = subprocess.run([executable, operation, mesh_file, mesh_file_rotated, os.path.join(path_folder_output, operation + "_" + mesh_file_output.replace(".off",""))], capture_output=True, text=True)

    #wait for the next step to be executed with a sleep of 3 minutes

    
    end_time = time.time() - start_time_float

    # Convert elapsed time to hours, minutes, and seconds
    hours = int(end_time // 3600)
    minutes = int((end_time % 3600) // 60)
    seconds = int(end_time % 60)
    milliseconds = int((end_time - int(end_time)) * 1000)

    # Format the elapsed time
    end_time_formatted = "{:02d}:{:02d}:{:02d}:{:02d}".format(hours, minutes, seconds, milliseconds)

    #remove from the mesh_file the path and the off extension

    mesh_file = mesh_file.replace(path_folder + "/", "")
    #mesh_file = mesh_file.replace("./data/test", "")
    if mesh_file.endswith(".off"):
        mesh_file = mesh_file.replace(".off", "")
    elif mesh_file.endswith(".stl"):
        mesh_file = mesh_file.replace(".stl", "")
    elif mesh_file.endswith(".obj"):
        mesh_file = mesh_file.replace(".obj", "")
   
    #check if the new file is in the results folder
    #pick the last file in the results folder
    
    files_output = os.listdir(path_folder_output)


    #remove the path and the extension from the mesh_file
    mesh_file_name = mesh_file.replace(".off", "")
    mesh_file_name = mesh_file_name.replace(path_folder_filename + "/", "") 

    mesh_file_name_rotated = mesh_file_rotated.replace(".off", "")
    mesh_file_name_rotated = mesh_file_name_rotated.replace(path_folder_filename + "/", "")





    if run_process.stderr:
        #print("Failed on file:", mesh_file, run_process.stderr)
        result = [mesh_file_name, mesh_file_name_rotated, operation, "Failed", end_time_formatted, run_process.stderr]
    elif not operation + "_"+ mesh_file_output.replace(".off","") in files_output:
        result = [mesh_file_name, mesh_file_name_rotated, operation, "Failed", end_time_formatted, "The output file is not in the results folder. Somenthing went wrong."]
    else: 
        result = [mesh_file_name, mesh_file_name_rotated, operation, "Passed", end_time_formatted, ""]
    

    write_to_excel(result, excel_file)


def create_excel_file(excel_file):

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Test Results"
    column_headers = ["File Tested 1", "File Tested 2", "Operation", "Run Status", "Completation Time (hh:mm:ss:ms)","Error Message"]

    # Writing headers
    for col_index, value in enumerate(column_headers, start=2):
        sheet.cell(row=2, column=col_index).value = value
        sheet.cell(row=2, column=col_index).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
        sheet.cell(row=1, column=col_index).font = openpyxl.styles.Font(bold=True)

    #center the text in the cells 3 and 4
    for i in range(3, len(column_headers)+2):
        sheet.cell(row=2, column=i).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)


    wb.save(excel_file)


def write_to_excel(output_data, excel_file):
    wb_add = load_workbook(excel_file)

    
    sheet = wb_add["Test Results"]
    ins_row = str(len(sheet['B']) + 1 )
    ins_row = int(ins_row)
    row_index = ins_row
   
    for col_index, value in enumerate(output_data, start=2):
        sheet.cell(row=row_index, column=col_index).value = value
        sheet.column_dimensions[sheet.cell(row=row_index, column=col_index).column_letter].auto_size = True
        sheet.cell(row=row_index, column=col_index).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
        if row_index != 2 and col_index == 5 : #if the value is "Failed" make the row red from the second column to the last
            if value == "Failed":
                for i in range(2, len(output_data)+2):
                    sheet.cell(row=row_index, column=i).fill = openpyxl.styles.PatternFill(start_color="FF4F47", end_color="FF4F47", fill_type = "solid")
                    sheet.cell(row=row_index, column=i).font = openpyxl.styles.Font(color="FFFFFF")
            else:
                for i in range(2, len(output_data)+2):
                    sheet.cell(row=row_index, column=i).fill = openpyxl.styles.PatternFill(start_color="D1FFBD", end_color="D1FFBD", fill_type = "solid")


    # Resize each column to fit the content
    for column_cells in sheet.columns:
        max_length = 0
        for cell in column_cells:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width


    #center the text in the cells 3 and 4
    for i in range(3, len(output_data)+2):
        sheet.cell(row=row_index, column=i).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Save the Excel file
    wb_add.save(excel_file)

import smtplib
from email.message import EmailMessage

def send_email_with_attachment(filename, content):
    # Set up email account information
    email_address = "sender.uni@hotmail.com"
    email_password = "Haloreach1!"
    smtp_server = "smtp-mail.outlook.com"
    smtp_port = 587

    # Create email message
    msg = EmailMessage()
    msg['Subject'] = 'Results from the server test run'
    msg['From'] = email_address
    msg['To'] = 'michele.faeddal@hotmail.com'
    msg.set_content(content)
    with open(filename, 'rb') as f:
        file_data = f.read()
    msg.add_attachment(file_data, maintype='application', subtype='octet-stream', filename=filename)

    # Send email message
    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            # Check if already logged in
            if not server.does_esmtp:
                server.starttls()
                server.login(email_address, email_password)
            server.send_message(msg)
        print("Email sent successfully!")
    except Exception as e:
        print("An error occurred:", e)



if __name__ == "__main__":

    path_folder = "./mesh_ready_to_test"
    path_folder_rotated ="./mesh_ready_to_test/mesh_rotated"
    folder_results = "./results_tests"
    excel_file = "test_results_mesh_booleans.xlsx"
    mode = "debug"
    operations = ["intersection", "union", "subtraction"]
    sendmail = False

    os.system('cls' if os.name == 'nt' else 'clear')
    if not os.path.exists(folder_results):
        meshes_dir = os.path.join(folder_results, 'meshes')
        os.makedirs(meshes_dir, exist_ok=True)

    
    print(os.path.join(folder_results, "meshes"))
    files = sorted_alphanumeric(os.listdir(path_folder))
    files_rotated = sorted_alphanumeric(os.listdir(path_folder_rotated))

    total_files = len(files)
    total_files_rotated = len(files_rotated)
    
    #create a for to loop over the files in pair and run the test
    for j in range(0, len(operations)):
        excel_file = "test_results_mesh_booleans_" + operations[j] + ".xlsx"
        if os.path.exists(os.path.join(folder_results, "meshes")):
            clean_folder(os.path.join(folder_results, "meshes"))


        #if the file exists, delete it
        if os.path.exists(os.path.join(folder_results, excel_file)):
            os.remove(os.path.join(folder_results, excel_file))

        #if the file does not exist, create it
        if not os.path.exists(os.path.join(folder_results, excel_file)):
            create_excel_file(os.path.join(folder_results, excel_file))

        for i, filename in enumerate(files):
            if filename.endswith(".off") or filename.endswith(".stl") or filename.endswith(".obj"):

                #find the filename + _rotaded in the rotated files list
                

                file_to_find = filename.replace(".off", "")

                
                filename_rotated =  [string for string in files_rotated if file_to_find +"_rotated.off" in string]
                filename_rotated = filename_rotated[0]
                


                filename_cleaned = filename.replace("_rotated", ""); 
                if filename_cleaned != filename_rotated.replace("_rotated", ""):
                    print("The files are not the same. Check the files in the folder", filename, filename_rotated)
                    exit()


                progress_bar = completion_bar(i, total_files-1)
                print("\r",operations[j].upper(),"Progress: {}% {}".format(int(((i+1) / total_files) * 100), progress_bar), "Files Processed -> ", i+1,"/", total_files," Running on file: ", filename, end="")
            
                mesh_file = os.path.join(path_folder, filename)
                mesh_file_rotated = os.path.join(path_folder_rotated, filename_rotated)


                result = compile_and_run_cpp(mesh_file, mesh_file_rotated, path_folder, operations[j], os.path.join(folder_results, excel_file), mode)
                
            
                # Print progress after each 1/5 of total files processed
                if sendmail and i % (total_files // 10) == 0:
                    progress = i // (total_files // 10)
                    send_email_with_attachment(excel_file,"\n"+operations[j].upper() + " - Progress: {}%".format(progress * 10))
        if sendmail:
            send_email_with_attachment(excel_file,"\n"+operations[j].upper() + " - Progress: {}%".format(100))




